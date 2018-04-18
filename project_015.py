import json

import openpyxl

with open('city.txt','r') as f:
    lines = f.readlines()
    text = "".join(lines)
    text = json.loads(text)
workbook = openpyxl.Workbook()
sheed = workbook.create_sheet('city')
row = 1
col = 1
for k,v in text.items():
    sheed.cell(row,col,value=k)
    sheed.cell(row, col+1, value=v)
    row += 1
workbook.save('city.xlsx')