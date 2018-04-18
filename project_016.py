import openpyxl

with open('numbers.txt','r') as f:
    lines = f.readlines()
    text = "".join(lines)
    list = eval(text)
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('numbers')
for i in range(len(list)):
    for j in range(len(list[i])):
        sheet.cell(i+1,j+1,value=list[i][j])
workbook.save('numbers.xlsx')